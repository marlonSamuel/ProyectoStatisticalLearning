# -*- coding: utf-8 -*-
"""
Created on Wed Jun 17 14:02:59 2020

@author: senpai
"""

from sklearn.metrics import confusion_matrix
from sklearn.metrics import classification_report
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn import metrics
import pandas as pd

labels = ["no sobrevivio","sobrevivio"]

def confussionMatrix(y_real,y_pred):
    conf_matrix = confusion_matrix(y_real, y_pred)
    plt.figure(figsize=(12, 12))
    sns.heatmap(conf_matrix, xticklabels=labels, yticklabels=labels, annot=True, fmt="d");
    plt.title("Confusion matrix")
    plt.ylabel('True class')
    plt.xlabel('Predicted class')
    plt.show()
    print("verdaderos positivos:",conf_matrix[1][1])
    print("falsos positivos:",conf_matrix[0][1])
    print("verdaderos negativos:",conf_matrix[0][0])
    print("falsos negativos:",conf_matrix[1][0])
    r = classification_report(y_real, y_pred)
    print(r)
    
def getMetrics(y_real,y_pred):
    return {
        "Accuracy": metrics.accuracy_score(y_real, y_pred),
        "Precision": metrics.precision_score(y_real, y_pred),
        "Recall:": metrics.recall_score(y_real, y_pred),
        "F1_Score:":metrics.f1_score(y_real, y_pred)
    }

def getReport(y_test,y_train,y_pred_test, y_pred_train):
    m_test = getMetrics(y_test,y_pred_test)
    m_train = getMetrics(y_train,y_pred_train)
    confussionMatrix(y_test,y_pred_test)
    return {
        "metrics_train":m_train,
        "metrics_test":m_test
    }

def writeInExcel(df,file_name,string):
    import os.path
    from openpyxl import load_workbook
    file_name = 'bitacoras/'+file_name
    if os.path.isfile(file_name):
        book = load_workbook(file_name)
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        startrow = book['Sheet1'].max_row
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        
        writer.sheets['Sheet1'].cell(row=startrow+2, column=2).value = string
        for sheetname in writer.sheets:
            df.to_excel(writer,sheet_name='Sheet1', startrow=startrow+3, index = True,header= True)

    else:
        writer = pd.ExcelWriter(file_name)
        df.to_excel(writer, startrow=3, startcol=0)

        worksheet = writer.sheets['Sheet1']
        worksheet.write(1, 1, string)
    writer.save()
    writer.close()


